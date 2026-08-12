"""Build the client-facing Excel workbook.

Layout (per project decision — easy to change):
  Sheet 1 "Summary"      — run info, per-county table, new/changed/buy-box highlights
  Sheet 2 "All Auctions" — one master tab, county column, filters + freeze panes
  Sheet 3 "Removed"      — rows that disappeared since last run (cancelled/redeemed/sold)
  Sheet 4 "Issues"       — parse warnings, county errors, excluded foreclosure rows
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import AuctionRecord

FILL_NEW = PatternFill("solid", start_color="C6EFCE")       # green
FILL_CHANGED = PatternFill("solid", start_color="FFEB9C")   # yellow
FILL_REMOVED = PatternFill("solid", start_color="D9D9D9")   # grey
FILL_HEADER = PatternFill("solid", start_color="1F4E79")    # dark blue
FILL_MATCH = PatternFill("solid", start_color="DDEBF7")     # light blue
FONT_HEADER = Font(color="FFFFFF", bold=True)
THIN = Border(bottom=Side(style="thin", color="CCCCCC"))

MASTER_COLUMNS = [
    ("County", "county", 12),
    ("Sale Date", "sale_date", 11),
    ("Sale Time", "sale_time", 10),
    ("Parcel ID", "parcel_id", 20),
    ("Case #", "case_number", 14),
    ("Certificate #", "certificate_number", 13),
    ("Owner Name", "owner_name", 22),
    ("Property Address", "property_address", 38),
    ("Property Use", "property_use", 18),
    ("Acreage", "acreage", 9),
    ("Opening Bid", "opening_bid", 13),
    ("Assessed Value", "assessed_value", 14),
    ("Bid/Assessed", "_bid_ratio", 12),
    ("Buy-Box", "_buybox", 9),
    ("Buy-Box Reasons", "_buybox_reasons", 30),
    ("Status", "_status", 11),
    ("Changes", "_changes", 30),
    ("Anomalies", "_anomalies", 26),
    ("Auction URL", "auction_url", 16),
    ("Appraiser URL", "appraiser_url", 16),
]


def _set_headers(ws, headers_widths):
    for i, (title, _, width) in enumerate(headers_widths, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers_widths))}1"


def build_workbook(out_path: str | Path, records: list[AuctionRecord], annotations: dict,
                   diff: dict, run_meta: dict, excluded: list[dict]) -> Path:
    """annotations: key() -> {buybox, buybox_reasons, anomalies}"""
    wb = Workbook()
    status_of = diff["status"]
    changes_of = diff["changes"]

    # ---- All Auctions (master tab) ----
    ws = wb.create_sheet("All Auctions")
    _set_headers(ws, MASTER_COLUMNS)
    records_sorted = sorted(records, key=lambda r: (r.sale_date[6:] + r.sale_date[:2] + r.sale_date[3:5], r.county, r.parcel_id))
    for ridx, rec in enumerate(records_sorted, start=2):
        ann = annotations.get(rec.key(), {})
        st = status_of.get(rec.key(), "")
        ratio = ""
        if rec.opening_bid and rec.assessed_value:
            ratio = round(rec.opening_bid / rec.assessed_value, 3)
        values = {
            "_bid_ratio": ratio,
            "_buybox": ann.get("buybox", ""),
            "_buybox_reasons": ann.get("buybox_reasons", ""),
            "_status": st,
            "_changes": "; ".join(changes_of.get(rec.key(), [])),
            "_anomalies": "; ".join(ann.get("anomalies", [])),
        }
        for cidx, (_, attr, _) in enumerate(MASTER_COLUMNS, start=1):
            val = values[attr] if attr in values else getattr(rec, attr, "")
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.border = THIN
            if attr in ("opening_bid", "assessed_value") and isinstance(val, (int, float)):
                cell.number_format = "$#,##0"
            if attr == "_bid_ratio" and isinstance(val, float):
                cell.number_format = "0.0%"
            if attr in ("auction_url", "appraiser_url") and val:
                cell.hyperlink = val
                cell.value = "link"
                cell.font = Font(color="0563C1", underline="single")
        fill = {"NEW": FILL_NEW, "CHANGED": FILL_CHANGED}.get(st)
        if fill:
            for cidx in range(1, len(MASTER_COLUMNS) + 1):
                ws.cell(row=ridx, column=cidx).fill = fill
        if ann.get("buybox") == "MATCH":
            ws.cell(row=ridx, column=14).font = Font(bold=True)
            ws.cell(row=ridx, column=14).fill = FILL_MATCH

    # ---- Removed ----
    ws_rm = wb.create_sheet("Removed")
    rm_cols = MASTER_COLUMNS[:12]
    _set_headers(ws_rm, rm_cols)
    for ridx, rec in enumerate(diff["removed"], start=2):
        for cidx, (_, attr, _) in enumerate(rm_cols, start=1):
            c = ws_rm.cell(row=ridx, column=cidx, value=getattr(rec, attr, ""))
            c.fill = FILL_REMOVED

    # ---- Issues ----
    ws_is = wb.create_sheet("Issues")
    _set_headers(ws_is, [("County", "", 12), ("Type", "", 18), ("Detail", "", 100)])
    ridx = 2
    for slug, centry in (run_meta.get("counties") or {}).items():
        if centry.get("status") == "error":
            ws_is.cell(row=ridx, column=1, value=slug)
            ws_is.cell(row=ridx, column=2, value="county error (skipped)")
            ws_is.cell(row=ridx, column=3, value=centry.get("error", ""))
            ridx += 1
        if centry.get("status") == "skipped_robots":
            ws_is.cell(row=ridx, column=1, value=slug)
            ws_is.cell(row=ridx, column=2, value="skipped by robots.txt")
            ws_is.cell(row=ridx, column=3, value=centry.get("robots", ""))
            ridx += 1
        for w in centry.get("warnings", []):
            ws_is.cell(row=ridx, column=1, value=slug)
            ws_is.cell(row=ridx, column=2, value="parse warning")
            ws_is.cell(row=ridx, column=3, value=w)
            ridx += 1
    for ex in excluded:
        ws_is.cell(row=ridx, column=1, value=ex.get("county", ""))
        ws_is.cell(row=ridx, column=2, value="EXCLUDED: foreclosure data")
        ws_is.cell(row=ridx, column=3,
                   value=f"{ex.get('reason','')} | parcel={ex.get('parcel_id','')} case={ex.get('case_number','')} type={ex.get('auction_type','')}")
        ridx += 1

    # ---- Summary (built last, placed first) ----
    ws_s = wb.active
    ws_s.title = "Summary"
    ws_s.column_dimensions["A"].width = 26
    for col, w in (("B", 14), ("C", 14), ("D", 14), ("E", 16), ("F", 14), ("G", 40)):
        ws_s.column_dimensions[col].width = w

    def put(row, col, value, bold=False, fill=None):
        c = ws_s.cell(row=row, column=col, value=value)
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill
        return c

    put(1, 1, "MADD Assets — Florida Tax Deed Scrub", bold=True)
    ws_s.cell(row=1, column=1).font = Font(bold=True, size=14)
    put(2, 1, "Generated")
    put(2, 2, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    put(3, 1, "Baseline comparison" if diff.get("baseline") else "First run (no previous run to compare)")

    n_new = sum(1 for s in status_of.values() if s == "NEW")
    n_chg = sum(1 for s in status_of.values() if s == "CHANGED")
    n_match = sum(1 for a in annotations.values() if a.get("buybox") == "MATCH")
    n_review = sum(1 for a in annotations.values() if a.get("buybox") == "REVIEW")
    row = 5
    put(row, 1, "Totals", bold=True); row += 1
    for label, val, fill in [
        ("Auctions (all counties)", len(records), None),
        ("NEW since last run", n_new, FILL_NEW),
        ("CHANGED since last run", n_chg, FILL_CHANGED),
        ("REMOVED since last run", len(diff["removed"]), FILL_REMOVED),
        ("Buy-box MATCH", n_match, FILL_MATCH),
        ("Buy-box REVIEW (verify use)", n_review, None),
        ("Excluded foreclosure rows", len(excluded), None),
    ]:
        put(row, 1, label)
        put(row, 2, val, bold=True, fill=fill)
        row += 1

    row += 1
    put(row, 1, "By county", bold=True); row += 1
    headers = ["County", "Auctions", "New", "Changed", "Buy-box", "Status", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = put(row, i, h, bold=True); c.fill = FILL_HEADER; c.font = FONT_HEADER
    row += 1
    by_county: dict[str, dict] = {}
    for rec in records:
        d = by_county.setdefault(rec.county, {"n": 0, "new": 0, "chg": 0, "match": 0})
        d["n"] += 1
        st = status_of.get(rec.key(), "")
        d["new"] += st == "NEW"
        d["chg"] += st == "CHANGED"
        d["match"] += annotations.get(rec.key(), {}).get("buybox") == "MATCH"
    meta_counties = run_meta.get("counties") or {}
    for slug in sorted(set(by_county) | set(meta_counties)):
        d = by_county.get(slug, {"n": 0, "new": 0, "chg": 0, "match": 0})
        centry = meta_counties.get(slug, {})
        put(row, 1, slug)
        put(row, 2, d["n"]); put(row, 3, d["new"]); put(row, 4, d["chg"]); put(row, 5, d["match"])
        put(row, 6, centry.get("status", ""))
        notes = centry.get("error") or "; ".join(centry.get("warnings", [])[:2])
        put(row, 7, notes)
        row += 1

    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def load_json(path: Path, default):
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default
